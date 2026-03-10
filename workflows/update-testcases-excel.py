"""
Test Case Excel Manager -- creates/updates workflows/testing/testcases.xlsx

Usage:
    python update-testcases-excel.py add --req-id REQ-001 --task "Fix barcode" --priority Medium \
        --preconditions "App running" --steps "1. Open app\n2. Upload file" \
        --expected "File processed" --edge-cases "Empty file\nCorrupt file" [--known-issues "..."]

    python update-testcases-excel.py update --tc-id TC-003 --status Pass
    python update-testcases-excel.py update --tc-id TC-003 --status Fail

    python update-testcases-excel.py last-id
"""

import argparse
import os
import sys
from datetime import date

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "testing", "testcases.xlsx")

HEADERS = [
    "TC ID", "REQ ID", "Task", "Date", "Priority",
    "Preconditions", "Steps to Test", "Expected Result",
    "Edge Cases", "Status", "Tester Notes", "Pass/Fail Date"
]

# Styles
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
LIGHT_BLUE_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
TOP_ALIGN = Alignment(vertical="top")

# Column widths
COL_WIDTHS = {
    1: 10,   # TC ID
    2: 10,   # REQ ID
    3: 30,   # Task
    4: 12,   # Date
    5: 10,   # Priority
    6: 30,   # Preconditions
    7: 45,   # Steps to Test
    8: 30,   # Expected Result
    9: 30,   # Edge Cases
    10: 14,  # Status
    11: 25,  # Tester Notes
    12: 14,  # Pass/Fail Date
}

STATUS_FILLS = {
    "Pending Test": YELLOW_FILL,
    "Pass": GREEN_FILL,
    "Fail": RED_FILL,
    "Reopen": RED_FILL,
}


def ensure_dir():
    testing_dir = os.path.dirname(EXCEL_PATH)
    if not os.path.exists(testing_dir):
        os.makedirs(testing_dir, exist_ok=True)


def create_workbook():
    """Create a new workbook with headers and formatting."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    for col_idx, width in COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"

    dv = DataValidation(
        type="list",
        formula1='"Pending Test,Pass,Fail,Reopen"',
        allow_blank=True
    )
    dv.error = "Please select: Pending Test, Pass, Fail, or Reopen"
    dv.errorTitle = "Invalid Status"
    ws.add_data_validation(dv)
    dv.add(f"J2:J1048576")

    return wb


def load_or_create():
    """Load existing workbook or create new one."""
    ensure_dir()
    if os.path.exists(EXCEL_PATH):
        return load_workbook(EXCEL_PATH)
    return create_workbook()


def _scan_md_max_tc(testing_dir):
    """Scan all test case markdown files for the highest TC-NNN reference."""
    import re
    max_num = 0
    if not os.path.isdir(testing_dir):
        return max_num
    for fname in os.listdir(testing_dir):
        if not fname.endswith(".md"):
            continue
        fpath = os.path.join(testing_dir, fname)
        try:
            with open(fpath, "r", encoding="utf-8") as f:
                content = f.read()
            for m in re.finditer(r"TC-(\d+)", content):
                num = int(m.group(1))
                if num > max_num:
                    max_num = num
        except Exception:
            pass
    return max_num


def get_next_tc_id(ws):
    """Get next TC ID by reading Excel rows AND all markdown test case files."""
    max_num = 0
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val and str(val).startswith("TC-"):
            try:
                num = int(str(val).replace("TC-", ""))
                if num > max_num:
                    max_num = num
            except ValueError:
                pass
    testing_dir = os.path.dirname(EXCEL_PATH)
    md_max = _scan_md_max_tc(testing_dir)
    if md_max > max_num:
        max_num = md_max
    return f"TC-{max_num + 1:03d}"


def apply_row_style(ws, row_idx):
    """Apply alternating row colors and borders."""
    is_even = (row_idx % 2 == 0)
    base_fill = LIGHT_BLUE_FILL if is_even else WHITE_FILL

    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.border = THIN_BORDER

        if col_idx == 10:
            status = cell.value or ""
            cell.fill = STATUS_FILLS.get(status, base_fill)
        else:
            cell.fill = base_fill

        if col_idx in (6, 7, 8, 9):
            cell.alignment = WRAP_ALIGN
        else:
            cell.alignment = TOP_ALIGN


def add_test_case(args):
    """Add a new test case row."""
    wb = load_or_create()
    ws = wb.active

    tc_id = get_next_tc_id(ws)
    row_idx = ws.max_row + 1
    if row_idx == 2 and ws.cell(row=2, column=1).value is None:
        row_idx = 2

    edge_cases = args.edge_cases or ""
    if args.known_issues:
        if edge_cases:
            edge_cases += "\n\nKnown Issues:\n" + args.known_issues
        else:
            edge_cases = "Known Issues:\n" + args.known_issues

    values = [
        tc_id,
        args.req_id or "",
        args.task or "",
        date.today().strftime("%Y-%m-%d"),
        args.priority or "Medium",
        args.preconditions or "",
        args.steps or "",
        args.expected or "",
        edge_cases,
        "Pending Test",
        "",  # Tester Notes
        "",  # Pass/Fail Date
    ]

    for col_idx, val in enumerate(values, 1):
        ws.cell(row=row_idx, column=col_idx, value=val)

    apply_row_style(ws, row_idx)
    wb.save(EXCEL_PATH)
    print(tc_id)


def update_status(args):
    """Update status of an existing test case."""
    if not os.path.exists(EXCEL_PATH):
        print("ERROR: testcases.xlsx not found", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    found = False
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val and str(val).strip() == args.tc_id.strip():
            ws.cell(row=row, column=10, value=args.status)
            ws.cell(row=row, column=12, value=date.today().strftime("%Y-%m-%d"))
            if args.notes:
                ws.cell(row=row, column=11, value=args.notes)
            apply_row_style(ws, row)
            found = True
            break

    if not found:
        print(f"ERROR: {args.tc_id} not found", file=sys.stderr)
        sys.exit(1)

    wb.save(EXCEL_PATH)
    print(f"{args.tc_id} updated to {args.status}")


def last_id(args):
    """Print the last TC ID (or TC-000 if none)."""
    if not os.path.exists(EXCEL_PATH):
        print("TC-000")
        return
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    tc_id = get_next_tc_id(ws)
    num = int(tc_id.replace("TC-", "")) - 1
    if num < 0:
        num = 0
    print(f"TC-{num:03d}")


def find_tc(args):
    """Find TC ID by req ID and task title."""
    if not os.path.exists(EXCEL_PATH):
        print("")
        return
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    req_id = (args.req_id or "").strip()
    task = (args.task or "").strip().lower()

    for row in range(ws.max_row, 1, -1):
        row_req = str(ws.cell(row=row, column=2).value or "").strip()
        row_task = str(ws.cell(row=row, column=3).value or "").strip().lower()
        if row_req == req_id and row_task == task:
            print(ws.cell(row=row, column=1).value)
            return

    print("")


def main():
    parser = argparse.ArgumentParser(description="Test Case Excel Manager")
    sub = parser.add_subparsers(dest="command")

    add_p = sub.add_parser("add", help="Add a new test case")
    add_p.add_argument("--req-id", default="")
    add_p.add_argument("--task", default="")
    add_p.add_argument("--priority", default="Medium")
    add_p.add_argument("--preconditions", default="")
    add_p.add_argument("--steps", default="")
    add_p.add_argument("--expected", default="")
    add_p.add_argument("--edge-cases", default="")
    add_p.add_argument("--known-issues", default="")

    upd_p = sub.add_parser("update", help="Update test case status")
    upd_p.add_argument("--tc-id", required=True)
    upd_p.add_argument("--status", required=True, choices=["Pass", "Fail", "Reopen", "Pending Test"])
    upd_p.add_argument("--notes", default="")

    find_p = sub.add_parser("find", help="Find TC ID by req ID and task title")
    find_p.add_argument("--req-id", default="")
    find_p.add_argument("--task", default="")

    sub.add_parser("last-id", help="Print last TC ID")

    args = parser.parse_args()

    if args.command == "add":
        add_test_case(args)
    elif args.command == "update":
        update_status(args)
    elif args.command == "find":
        find_tc(args)
    elif args.command == "last-id":
        last_id(args)
    else:
        parser.print_help()
        sys.exit(1)


if __name__ == "__main__":
    main()
